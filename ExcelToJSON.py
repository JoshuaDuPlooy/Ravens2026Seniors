import tkinter as tk
from tkinter import filedialog, messagebox
import json
import os
from datetime import datetime, date, time

try:
    import openpyxl
except ImportError:
    messagebox.showerror("Missing package", "Please install openpyxl:\n\npip install openpyxl")
    raise


def format_value(v):
    """Convert Excel date/time values to clean strings; leave everything else as-is."""
    if isinstance(v, datetime):
        # datetime covers both date+time — format as date only ("30 May")
        return v.strftime("%#d %b")
    if isinstance(v, date):
        return v.strftime("%#d %b")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    return v


EVENTS = [
    {"label": "Men's Open",    "output": "MensKnockouts.json"},
    {"label": "Ladies Open",   "output": "LadiesKnockouts.json"},
    {"label": "2nd & Below",   "output": "2ndAndBelowKnockouts.json"},
    {"label": "5th & Below",   "output": "5thAndBelowKnockouts.json"},
]


class ExcelToJSONApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel → JSON Converter")
        self.root.resizable(False, False)

        self.file_vars = []   # StringVar for each file path
        self.status_vars = [] # StringVar for each row status
        self.out_dir_var = tk.StringVar(value=os.path.dirname(os.path.abspath(__file__)))

        self.build_ui()

    def build_ui(self):
        # ── Title ──
        tk.Label(
            self.root,
            text="Senior Open 2026 – Excel to JSON",
            font=("Arial", 13, "bold"),
            pady=10
        ).grid(row=0, column=0, columnspan=4, padx=16)

        tk.Label(
            self.root,
            text="Select an Excel file for each event, then click Convert All.",
            font=("Arial", 9),
            fg="#555"
        ).grid(row=1, column=0, columnspan=4, padx=16, pady=(0, 12))

        # ── Output folder row ──
        tk.Label(
            self.root,
            text="Output folder",
            font=("Arial", 10, "bold"),
            width=12,
            anchor="w"
        ).grid(row=2, column=0, padx=(16, 4), pady=(0, 8), sticky="w")

        tk.Entry(
            self.root,
            textvariable=self.out_dir_var,
            width=40,
            state="readonly"
        ).grid(row=2, column=1, padx=4, pady=(0, 8))

        tk.Button(
            self.root,
            text="Browse…",
            width=8,
            command=self.browse_output
        ).grid(row=2, column=2, padx=4, pady=(0, 8))

        tk.Frame(self.root, height=1, bg="#ccc").grid(
            row=3, column=0, columnspan=4, sticky="ew", padx=16, pady=(0, 8)
        )

        # ── File rows ──
        for i, event in enumerate(EVENTS):
            row = i + 4

            # Event label
            tk.Label(
                self.root,
                text=event["label"],
                font=("Arial", 10, "bold"),
                width=12,
                anchor="w"
            ).grid(row=row, column=0, padx=(16, 4), pady=4, sticky="w")

            # File path entry
            file_var = tk.StringVar()
            self.file_vars.append(file_var)
            tk.Entry(
                self.root,
                textvariable=file_var,
                width=40,
                state="readonly"
            ).grid(row=row, column=1, padx=4, pady=4)

            # Browse button
            tk.Button(
                self.root,
                text="Browse…",
                width=8,
                command=lambda v=file_var: self.browse(v)
            ).grid(row=row, column=2, padx=4, pady=4)

            # Status label
            status_var = tk.StringVar(value="")
            self.status_vars.append(status_var)
            tk.Label(
                self.root,
                textvariable=status_var,
                width=4,
                font=("Arial", 12)
            ).grid(row=row, column=3, padx=(4, 16), pady=4)

        # ── Divider ──
        tk.Frame(self.root, height=1, bg="#ccc").grid(
            row=8, column=0, columnspan=4, sticky="ew", padx=16, pady=8
        )

        # ── Convert All button ──
        tk.Button(
            self.root,
            text="Convert All",
            font=("Arial", 11, "bold"),
            width=16,
            command=self.convert_all
        ).grid(row=9, column=0, columnspan=4, pady=(0, 4))

        # ── Log area ──
        self.log = tk.Text(
            self.root,
            height=6,
            width=60,
            state="disabled",
            font=("Courier", 9),
            bg="#f9f9f9",
            relief="flat",
            bd=1,
            highlightthickness=1,
            highlightbackground="#ccc"
        )
        self.log.grid(row=10, column=0, columnspan=4, padx=16, pady=(4, 16), sticky="ew")

    def browse_output(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.out_dir_var.set(path)

    def browse(self, file_var):
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
        )
        if path:
            file_var.set(path)

    def log_write(self, msg):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.config(state="disabled")

    def convert_all(self):
        # Clear log and statuses
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")
        for sv in self.status_vars:
            sv.set("")

        out_dir = self.out_dir_var.get()
        if not out_dir or not os.path.isdir(out_dir):
            self.log_write("✗  Invalid output folder. Please select a valid folder.")
            return

        any_done = False
        for i, event in enumerate(EVENTS):
            path = self.file_vars[i].get()
            if not path:
                self.log_write(f"[{event['label']}] Skipped — no file selected.")
                self.status_vars[i].set("–")
                continue

            try:
                wb = openpyxl.load_workbook(path, data_only=True)
                sheet_name = "Knockout Matches"
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}")
                ws = wb[sheet_name]

                # First row = headers
                headers = [str(cell.value).strip() if cell.value is not None else f"Col{j+1}"
                           for j, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))]

                records = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Skip completely empty rows
                    if all(v is None for v in row):
                        continue
                    record = {}
                    for h, v in zip(headers, row):
                        record[h] = format_value(v) if v is not None else ""
                    records.append(record)

                out_path = os.path.join(out_dir, event["output"])
                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(records, f, indent=2, ensure_ascii=False)

                self.log_write(f"[{event['label']}] ✓  {len(records)} rows → {event['output']}")
                self.status_vars[i].set("✓")
                any_done = True

            except Exception as e:
                self.log_write(f"[{event['label']}] ✗  Error: {e}")
                self.status_vars[i].set("✗")

        if any_done:
            self.log_write("\nDone! JSON files saved to the website folder.")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelToJSONApp(root)
    root.mainloop()
